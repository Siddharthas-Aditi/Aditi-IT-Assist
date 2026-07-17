import io
from datetime import UTC, datetime

from openpyxl import load_workbook

from app.schemas.reporting import SpecialistReport, SpecialistReportRow
from app.services.reporting import exporters


def _report():
    row = SpecialistReportRow(
        agent_id="a1",
        agent_name="Alex Agent",
        agent_email="alex@aditi.com",
        total_tickets=10,
        reopened=1,
        avg_resolution_hours=2.5,
        sla_violations=0,
        csat_avg=4.8,
        dsat=0,
        feedback_responses=6,
    )
    empty = SpecialistReportRow(
        agent_id="a2",
        agent_name="Blank Agent",
        total_tickets=0,
        csat_avg=None,
        avg_resolution_hours=None,
    )
    totals = SpecialistReportRow(
        agent_id=None,
        agent_name="Team totals",
        total_tickets=10,
        reopened=1,
        avg_resolution_hours=2.5,
        sla_violations=0,
        csat_avg=4.8,
        dsat=0,
        feedback_responses=6,
    )
    return SpecialistReport(
        period_start=datetime(2026, 7, 1, tzinfo=UTC),
        period_end=datetime(2026, 7, 31, tzinfo=UTC),
        rows=[row, empty],
        totals=totals,
    )


def test_csv_has_header_rows_and_dash_for_none():
    data = exporters.to_csv(_report()).decode("utf-8")
    assert "Agent" in data and "Alex Agent" in data and "Team totals" in data
    assert "-" in data  # None rendered as dash for Blank Agent's csat


def test_xlsx_opens_and_has_rows():
    wb = load_workbook(io.BytesIO(exporters.to_xlsx(_report())))
    ws = wb.active
    values = [c.value for c in next(ws.iter_rows())]
    assert "Agent" in values


def test_pdf_starts_with_magic():
    assert exporters.to_pdf(_report())[:4] == b"%PDF"


def _malicious_report():
    row = SpecialistReportRow(
        agent_id="a3",
        agent_name="=cmd|' /c calc'!A1",
        agent_email="evil@aditi.com",
        total_tickets=1,
        reopened=0,
        avg_resolution_hours=1.0,
        sla_violations=0,
        csat_avg=5.0,
        dsat=0,
        feedback_responses=1,
    )
    totals = SpecialistReportRow(
        agent_id=None,
        agent_name="Team totals",
        total_tickets=1,
        reopened=0,
        avg_resolution_hours=1.0,
        sla_violations=0,
        csat_avg=5.0,
        dsat=0,
        feedback_responses=1,
    )
    return SpecialistReport(
        period_start=datetime(2026, 7, 1, tzinfo=UTC),
        period_end=datetime(2026, 7, 31, tzinfo=UTC),
        rows=[row],
        totals=totals,
    )


def test_csv_neutralizes_formula_injection():
    data = exporters.to_csv(_malicious_report()).decode("utf-8")
    assert "'=cmd|' /c calc'!A1" in data
    for line in data.splitlines():
        for cell in line.split(","):
            assert not cell.startswith('"=') and not cell.startswith("=")


def test_csv_leaves_safe_names_unchanged():
    data = exporters.to_csv(_report()).decode("utf-8")
    assert "Alex Agent" in data
    assert "'Alex Agent" not in data


def test_xlsx_neutralizes_formula_injection():
    wb = load_workbook(io.BytesIO(exporters.to_xlsx(_malicious_report())))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    agent_name = rows[1][0]
    assert agent_name.startswith("'")
    assert not agent_name.startswith("=")
    assert agent_name == "'=cmd|' /c calc'!A1"


def test_xlsx_leaves_safe_names_unchanged():
    wb = load_workbook(io.BytesIO(exporters.to_xlsx(_report())))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    agent_name = rows[1][0]
    assert agent_name == "Alex Agent"
